import gym
from matplotlib import pyplot as plt
from torch.autograd import Variable
from collections import namedtuple
import torch
import torch.nn as nn
import random
import time
import numpy as np
import torch.nn.functional as F
from highway_env import utils
import highway_env
env = gym.make('highway-action-perturbation-v0').unwrapped
import sys

import warnings
warnings.filterwarnings("ignore")
import openpyxl as op

name = "oar_new_20w_dqn"

# 设置随机种子
seed_exp = 2
import random
random.seed(seed_exp)

import numpy as np
np.random.seed(seed_exp)

import torch
torch.manual_seed(seed_exp)
if torch.cuda.is_available():
    torch.cuda.manual_seed_all(seed_exp)

# 检查是否可以使用GPU
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
print(device)

from DQN_HEAD.DQN_HEAD import get_config
parser = get_config()
config = parser.parse_args()
###################   Basic   ##########################
env.config["lanes_count"] = config.lanes_count  # 4
env.config["show_trajectories"] = config.show_trajectories
env.config["screen_height"] = config.screen_height  # 150
env.config["screen_width"] = config.screen_width  # 600
env.config["reward_speed_range"] = config.reward_speed_range     #   设定速度范围 [2, 4, 6, 8]

env.config["policy_frequency"] = config.policy_frequency  #  abstract： frames = int(self.config["simulation_frequency"] // self.config["policy_frequency"])
#############  DQN 调参  ###########################
env.config["simulation_frequency"] = config.simulation_frequency  # 决定dt=1/"simulation_frequency"；即：一次step()，系统做出几次控制循环。还和 controller 比例控制 TAU_ACC = 2 有关，一并调整
# 太大，控制结束了，但是下一个动作 a 还没有来。太小的话，控制还没完，下一个动作 a 就来了# KP_A = 1 / 0.375 = 8/3   dt * KP_A = 8/3 * 1/4 = 2/3
env.config["vehicles_count"] = config.vehicles_count #  旁车，数据越大，训练频率变慢，显示越稳定. 每次控制循环实施的次数："vehicles_count"+1
env.config["duration"] = config.duration  # 40  决定了在不 crashed 的前提下，能有多少个 a 和 env.step(a)。直到steps计数到["duration"]，触发done，系统复位。

env.config["collision_reward"] = config.collision_reward  # -1
env.config["high_speed_reward"] = config.high_speed_reward  # 0.4
env.config["punish_speed_reward"] = config.punish_speed_reward
env.config["wrong_act_reward"] = config.wrong_act_reward  #
env.config["correct_act_reward"] = config.correct_act_reward

T_N_update_FQ = config.T_N_update_FQ  # 更新目标网络频率
BATCH_SIZE = config.BATCH_SIZE  # 提取批次数量  20
MEMORY_CAPACITY = config.MEMORY_CAP * BATCH_SIZE        # 池子容量  200 * 20

total_train_times = config.total_train_t * MEMORY_CAPACITY    # 训练次数  25 * 4000 = 10W

class DQNNet(nn.Module):
    def __init__(self):
        super(DQNNet, self).__init__()
        self.linear1 = nn.Linear(config.input_net, config.neure)
        self.linear1.weight.data.normal_(0, 0.1)
        self.linear2 = nn.Linear(config.neure, config.neure)
        self.linear2.weight.data.normal_(0, 0.1)
        self.linear3 = nn.Linear(config.neure, config.output_net)
        self.linear3.weight.data.normal_(0, 0.1)

    def forward(self, s):
        if not isinstance(s, torch.Tensor):
            s = torch.as_tensor(s, dtype=torch.float32, device=device)
        else:
            s = s.to(device)
        s01 = s.view(s.size(0), 1, 20)  # size(0)返回s的行数，view把s拉成1个1行25列才能送入网络    改！
        s02 = self.linear1(s01)
        s03 = F.relu(s02)
        s04 = self.linear2(s03)
        out = F.relu(s04)  # out = torch.sigmoid(s02)
        act_value = self.linear3(out)
        return act_value  # 1*5

class DQN(object):
    def __init__(self):
        self.eval_net, self.target_net = DQNNet().to(device), DQNNet().to(device)
        self.learn_step_counter = 0
        self.memory = []
        self.position = 0
        self.capacity = MEMORY_CAPACITY
        self.optimizer = torch.optim.Adam(self.eval_net.parameters(), lr=config.LR)
        self.loss_func = nn.MSELoss()
        self.index = 0
        for param, target_param in zip(self.eval_net.parameters(), self.target_net.parameters()):
            target_param.data.copy_(param.data)
        self.ZHEKOU = 0.01

    def choose_action(self, s, e):
        # 将状态s转换为torch变量，并设置requires_grad=True以便计算梯度
        x = np.expand_dims(s, axis=0)  # 增加了一维，多了一对中括号
        x = torch.tensor(x, dtype=torch.float32, device=device,
                         requires_grad=True)  # 使用torch.tensor并设置requires_grad=True

        if np.random.uniform() < 1 - e:  # 根据e选择随机动作还是根据策略选择动作
            actions_value = self.eval_net.forward(x)
            action = torch.max(actions_value, -1)[1].item()

            # 计算损失
            loss = -actions_value.max()  # 注意：这里我们最大化动作值，所以在反向传播时需要对其取负
            self.eval_net.zero_grad()
            loss.backward()

            # 应用FGSM攻击：基于梯度的方向生成对抗样本
            epsilon = 0.05  # 扰动程度
            # 注意：确保x.grad存在且不为None
            if x.grad is not None:
                x_adv = x.detach() + epsilon * x.grad.data.sign()
            else:
                x_adv = x.detach()  # 如果没有梯度信息，则不对输入做修改

            # 使用对抗样本来代替原始样本做决策
            actions_value_adv = self.eval_net.forward(x_adv)
            action = torch.max(actions_value_adv, -1)[1].item()
        else:
            action = np.random.randint(0, 5)

        return action

    def push_memory(self, s, a, r, s_):
        if len(self.memory) < self.capacity:  # MEMORY_CAPACITY = self.capacity
            self.memory.append(None)
        self.index = self.position % self.capacity
        self.memory[self.index] = Transition(torch.unsqueeze(torch.FloatTensor(s).to(device), 0),
                                             torch.unsqueeze(torch.FloatTensor(s_).to(device), 0), \
                                             torch.from_numpy(np.array([a])).to(device),
                                             torch.from_numpy(np.array([r], dtype='float32')).to(device))
        self.position += 1
        # print('total com times:', self.position )

    def learn(self):
        # if self.learn_step_counter % TARGET_NETWORK_REPLACE_FREQ == 0:
        #     self.target_net.load_state_dict(self.eval_net.state_dict())
        transitions = random.sample(self.memory, BATCH_SIZE)
        batch = Transition(*zip(*transitions))  # 按照类型重建元组，s, s_, a, r, 每个元素里面分别有 BATCH_SIZE 个
        b_s = torch.cat(batch.state)
        b_s_ = Variable(torch.cat(batch.next_state))
        b_a = Variable(torch.cat(batch.action))
        b_r = Variable(torch.cat(batch.reward))

        q_eval = self.eval_net.forward(b_s).squeeze(1).gather(1, b_a.unsqueeze(1).to(torch.int64))
        q_next = self.target_net.forward(b_s_).squeeze(1).detach()
        q_target = b_r + config.GAMMA * q_next.max(1)[0].view(BATCH_SIZE, 1).t()
        loss = self.loss_func(q_eval, q_target.t())
        # print('loss:', loss)
        self.optimizer.zero_grad()  # reset the gradient to zero
        loss.backward()
        self.optimizer.step()  # execute back propagation for one step

        if self.learn_step_counter % T_N_update_FQ == 0:  # 一开始触发，然后每100步触发
            for param, target_param in zip(self.eval_net.parameters(), self.target_net.parameters()):
                target_param.data.copy_((1 - self.ZHEKOU) * target_param.data + self.ZHEKOU * param.data)

        self.learn_step_counter += 1

        return loss.detach().cpu().numpy()

    def save_model(self, path):
        """保存模型"""
        torch.save(self.eval_net.state_dict(), path)

    def load_model(self, path):
        """加载模型"""
        self.eval_net.load_state_dict(torch.load(path))
        self.target_net.load_state_dict(self.eval_net.state_dict())  # 确保目标网络也同步加载

dqn = DQN()

Transition = namedtuple('Transition', ('state', 'next_state', 'action', 'reward'))  # namedtuple()是继承自tuple的子类，namedtuple()方法能够创建一个和tuple类似的对象，而且对象拥有可访问的属性

wb = op.Workbook()  # 创建工作薄对象
ws = wb['Sheet']  # 创建子表

#model_path = 'model_fine_tuned_dqn/DQN_20250109_023208.pth'
#model = dqn.load_model(model_path)

while True:  # 原始程序
    done = False
    s = env.reset()   # 5*4, r, false, {speed, crashed, action, cost}
    config.episode += 1
    config.S_tm_ep = time.time()
    while not done:
        #env.render()  # 显示动画，真的没必要,大大降低程序运行速度。        # e = np.exp(-total_env_times / MEMORY_CAPACITY)
        a = dqn.choose_action(s, config.e)   # 对应 ACTIONS_ALL = { 0: 'LANE_LEFT', 1: 'IDLE', 2: 'LANE_RIGHT',  3: 'FASTER',  4: 'SLOWER'   }
        config.total_step_times += 1

        if ( config.start_train_TR == 0):  # AR 训练前
            if ( config.lane_old == 0 and a == 0) or ( config.lane_old == env.config["lanes_count"]-1 and a == 2):  # 如果出现不合理动作，触发惩罚标记 WRO_ACT_TR
                config.WRO_ACT_TR = 1  # 此处标记是为了触发惩罚。 合理动作判断！
            else:  # 合理动作
                config.CORRECT_ACT_TR = 1

            s_, r, done, info, SUCCESS = env.step(a)  # 错误动作也进入训练，带惩罚。 success 表示成功完成全部轮次。env.config["duration"] = 90
            if SUCCESS == True:
                print("????????????????????????????????????????")
            config.lane_old = r[1]

            scaled_speed = utils.lmap(info['speed'], env.config["reward_speed_range"], [0, 1])  # "reward_speed_range"= [7,9];
            h_spd = env.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
            config.reward = r[0] + h_spd

            config.new_reward = config.reward + env.config["correct_act_reward"] * config.CORRECT_ACT_TR
            config.reward = utils.lmap(config.new_reward,
                                       [env.config["collision_reward"], env.config["correct_act_reward"] + env.config["high_speed_reward"]], [0, 1])

            #print(f"{config.episode}:{config.total_step_times} \n reward:{config.reward} \n new_reward:{config.new_reward} \n reward:{config.reward}")

            # reward perturbation
            config.max_reward = max(config.max_reward,config.reward)
            if config.reward >= config.max_reward:
                config.reward = 2*config.max_reward-config.reward - 0.2*config.max_reward

            config.WRO_ACT_TR = 0
            config.CORRECT_ACT_TR = 0

            if not done:  # 合理还正确，不撞车。     训练前 填池子，训练开始后池子数量不增加
                dqn.push_memory(s, a, config.reward, s_)
                config.memory_cnt += 1
            else:  # 合理不正确，动作引起的下一时刻撞车动作
                dqn.push_memory(s, a, config.reward, s_)  # 错误动作进池子
                config.memory_cnt += 1  # 错误动作进池子

                config.E_tm_ep = time.time()
                config.tm_ep = config.E_tm_ep - config.S_tm_ep
                config.tm_ep_array.append(config.tm_ep)
                config.ntm += 1
        #========================# 训练后 #===========================#
        elif (config.start_train_TR == 1):  # 训练后
            if (config.lane_old == 0 and a == 0) or (config.lane_old == env.config["lanes_count"]-1 and a == 2):  # 如果出现不合理动作，触发惩罚标记 WRO_ACT_TR
                config.WRONG_ACT_CNT_in_train += 1
                config.WRO_ACT_TR = 1
            else:   # 出现合理动作
                config.correct_act_cnt_in_train += 1
                config.CORRECT_ACT_TR = 1
                config.sum_RA_episode +=1

            s_, r, done, info, SUCCESS = env.step(a)  # info里面含有 speed， crushed，action 的信息。 如：if info['crashed'] == True :
            lane_old = r[1]

            scaled_speed = utils.lmap(info['speed'], env.config["reward_speed_range"], [0, 1])  # "reward_speed_range"= [3,9];
            h_spd = env.config["high_speed_reward"] * np.clip(scaled_speed, 0, 1)
            config.reward = r[0] + h_spd

            config.new_reward = config.reward + env.config["correct_act_reward"] * config.CORRECT_ACT_TR
            config.reward = utils.lmap(config.new_reward,
                                       [env.config["collision_reward"], env.config["correct_act_reward"] + env.config["high_speed_reward"]], [0, 1])

            # reward perturbation
            config.max_reward = max(config.max_reward,config.reward)
            if config.reward >= config.max_reward:
                config.reward = 2*config.max_reward-config.reward - 0.2*config.max_reward
            config.reward_array.append(config.reward) # 积累奖励

            config.WRO_ACT_TR = 0
            config.CORRECT_ACT_TR = 0

            if np.all(s_[1] == [0, 0, 0, 0]) & np.all(s_[2] == [0, 0, 0, 0]):   #  只记录完成避障所用时间
                config.E_tm_ep = time.time()
                config.tm_ep = config.E_tm_ep - config.S_tm_ep
                config.tm_ep_array.append(config.tm_ep)
                config.ntm += 1
                done = True    # 完成避障

            config.speed_array.append(info['speed'] )

            if not done:  # 训练前 填池子，训练开始后池子数量不增加
                dqn.push_memory(s, a, config.reward, s_)
            else:  # 有可能成功避障，如果是成功完成避障，后面记录第一次成功的时间
                dqn.push_memory(s, a, config.reward, s_)  # 错误动作进池子
                config.done_in_train += 1  # 正确和错误的动作都会引发下一个 done ！！！！！！！

                config.sum_reward_episode = np.sum(config.reward_array)  # 每一轮奖励和
                config.reward_sum_in_train += config.sum_reward_episode  # 全部训练奖励和 最后输出数据用
                config.sum_reward_episode_array.append(config.sum_reward_episode)  # 构成奖励数组 绘图用
                config.nreward += 1
                config.reward_array = []

                config.avg_speed_episode = np.mean(config.speed_array)  # 轮次平均速度
                config.avg_speed_episode_array.append(config.avg_speed_episode)  # 所有轮次平均速度 绘图用
                config.nspeed += 1
                config.speed_array = []

                config.sum_RA_episode_array.append(config.sum_RA_episode)  # 每一轮的RA数量进入数组
                config.nRA += 1
                config.sum_RA_episode = 0

                config.avg_loss_episode = np.mean(config.loss_array)
                config.avg_loss_episode_array.append(config.avg_loss_episode)    # 所有轮次平均损失 绘图用
                config.nloss += 1
                config.loss_array = []

                config.E_tm_ep = time.time()
                config.tm_ep = config.E_tm_ep - config.S_tm_ep
                config.tm_ep_array.append(config.tm_ep)
                config.ntm += 1

                if np.all(s_[1] == [0,0,0,0]) & np.all(s_[2] == [0,0,0,0]):   # 计算每一次完成避障时间
                    config.success_done += 1  # env.commom.abstract 中定义
                    if (config.success_done == 1):
                        end_time = time.time()  # 计算第一次成功避障所用时间
                        config.First_finish_time = end_time - start_time

        s = s_
        if config.total_step_times % 1000 == 0:
            print('*  DQN_C total_step_times:', config.total_step_times)
################################  开始训练  ###################################################
        if dqn.position >= MEMORY_CAPACITY:  # 池子满了，开始训练
            config.start_train_TR = 1
            config.trained_times += 1
            if (config.trained_times ==1 ):
                start_time = time.time()   # 训练时间计时
            loss = dqn.learn()
            config.loss_array.append(loss)

            e = 1/(config.trained_times/50 + 10)

            # 全部训练完成，显示结果。跳出循环
            if (config.trained_times == total_train_times + 1 ):
                end_time = time.time()
                total_time = end_time - start_time

                print('------------------------------------', )
                print('==== DQN_C Results ===============', )
                print('   memory_cnt:', config.memory_cnt)
                print('      训练次数:', config.trained_times - 1)
                print('   总体训练时间:', np.round(total_time, 0)  )
                print('   首次成功用时:', np.round(config.First_finish_time, 0)  )
                print('训练后合理动作数:', config.correct_act_cnt_in_train)
                print('   合理动作比例:', np.round(100 * config.correct_act_cnt_in_train / (config.trained_times - 1), 2), '%')
                print('   训练奖励总和:', np.round(config.reward_sum_in_train, 0))
                print('     训练后轮次：', config.done_in_train)
                print('   成功避障次数:', config.success_done)
                print('        成功率:', np.round(100 * config.success_done/(config.done_in_train+0.1), 2), '%')
                print('            e:', np.round(e, 6))
                print('-------------------------------', )
                print('Waiting for writing excel......')
                import numpy as np

                # 假设config和e等变量已经定义并赋值
                output = f"""------------------------------------
                ==== DQN_C Results ===============
                   memory_cnt: {config.memory_cnt}
                      训练次数: {config.trained_times - 1}
                   总体训练时间: {np.round(total_time, 0)}
                   首次成功用时: {np.round(config.First_finish_time, 0)}
                训练后合理动作数: {config.correct_act_cnt_in_train}
                   合理动作比例: {np.round(100 * config.correct_act_cnt_in_train / (config.trained_times - 1), 2)}%
                   训练奖励总和: {np.round(config.reward_sum_in_train, 0)}
                     训练后轮次： {config.done_in_train}
                   成功避障次数: {config.success_done}
                        成功率: {np.round(100 * config.success_done / (config.done_in_train + 0.1), 2)}%
                            e: {np.round(e, 6)}
                ------------------------------------
                Waiting for writing excel......
                """

                with open("exp_" + name + "_result.txt", 'w') as file:
                    file.write(output)

                # 保存数据到xlsx表

                import os
                if not os.path.exists("xlsx_" + name ):
                    os.makedirs("xlsx_" + name )

                for e in range(config.nreward):
                    ws.cell(row=e + 1, column=1).value = float(config.sum_reward_episode_array[e])  #
                    wb.save("xlsx_" + name + '/reward.xlsx')  # 保存excel表

                for i in range(config.nloss): # 一般不看loss
                    ws.cell(row=i + 1, column=1).value = float(config.avg_loss_episode_array[i])  # 在第1行第1列写入LOSS
                    wb.save("xlsx_" + name + '/loss.xlsx')  # 保存excel表

                for d in range(config.nspeed):
                    ws.cell(row=d + 1, column=1).value = float(config.avg_speed_episode_array[d])  #
                    wb.save("xlsx_" + name + '/speed.xlsx')  # 保存excel表

                for f in range(config.nRA):
                    ws.cell(row=f + 1, column=1).value = float(config.sum_RA_episode_array[f])  #
                    wb.save("xlsx_" + name + '/ra.xlsx')  # 保存excel表

                for g in range(config.ntm):
                    ws.cell(row=g + 1, column=1).value = float(config.tm_ep_array[g])  #
                    wb.save("xlsx_" + name + '/tm_ep.xlsx')  # 保存excel表

                print('Writing excel is finished.')

                # 保存训练好的模型
                from datetime import datetime

                # 获取当前日期时间和格式化
                formatted_date_time = datetime.now().strftime("%Y%m%d_%H%M%S")
                print("Formatted date and time:", formatted_date_time)

                if not os.path.exists("model_oar"):
                    os.makedirs("model_oar")

                model_path = "model_oar/" + name + "_" + formatted_date_time + ".pth"
                dqn.save_model(model_path)
                print(f"Model saved to {model_path}")

                sys.exit()

